from .Styles import styles
from openpyxl.utils.cell import get_column_letter

def create_sheet(work_book, platform):
    work_sheet = work_book.active
    work_sheet.title = platform
    return work_sheet


def set_size(work_sheet, data):
    work_sheet.row_dimensions[1].height = 30
    work_sheet.row_dimensions[2].height = 20

    for column_number in range(len(data.column_name_list)):
        width = len(data.column_name_list[column_number])
        for row_number in range(len(data.table)):
            if len(data.table[row_number][column_number]) > width:
                width = len(data.table[row_number][column_number])
        work_sheet.column_dimensions[get_column_letter(column_number + 1)].width = width + 10


def set_titles(work_sheet, data):
    work_sheet['A1'] = data.platform
    work_sheet.merge_cells(f'A1:{get_column_letter(len(data.column_name_list))}1')

    work_sheet.append(data.column_name_list)


def set_data(work_sheet, data):
    for table_string_list in data.table:
        work_sheet.append(table_string_list)


def set_styles(work_sheet, data):
    row_numbers_list = range(2, (len(data.table) + 3))
    column_numbers_list = range(1, (len(data.column_name_list) + 1))

    align = styles.alignment()
    work_sheet['A1'].alignment = align
    for row_number in row_numbers_list:
        for column_number in column_numbers_list:
            work_sheet.cell(row=row_number, column=column_number).alignment = align

    title_font = styles.title_font()
    work_sheet['A1'].font = title_font
    for column_number in column_numbers_list:
        work_sheet.cell(row=2, column=column_number).font = styles.title_font()

    border = styles.border('bottom')
    work_sheet['A1'].border = border
    for column_number in column_numbers_list:
        work_sheet.cell(row=1, column=column_number).border += border
        work_sheet.cell(row=2, column=column_number).border += border
        work_sheet.cell(row=(len(data.table) + 2), column=column_number).border += border

    border = styles.border('right')
    for column_number in column_numbers_list:
        work_sheet.cell(row=2, column=column_number).border += border
    for row_number in row_numbers_list:
        work_sheet.cell(row=row_number, column=len(data.column_name_list)).border += border
