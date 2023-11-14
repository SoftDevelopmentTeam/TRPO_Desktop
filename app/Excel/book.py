from openpyxl import Workbook as book
from .Sheets import sheets


def create_book():
    return book()


def save_book(work_book, path):
    work_book.save(path)


def create_sheet(work_book, data):
    work_sheet = sheets.create_sheet(work_book, data.platform)
    sheets.set_size(work_sheet, data)
    sheets.set_titles(work_sheet, data)
    sheets.set_data(work_sheet, data)
    sheets.set_styles(work_sheet, data)


def main(data, path):
    work_book = create_book()
    create_sheet(work_book, data)
    save_book(work_book, path)
