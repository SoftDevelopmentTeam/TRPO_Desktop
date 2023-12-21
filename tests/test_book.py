import pytest
import openpyxl
from app.Excel import book
from app.Parser import parser
import pathlib


pairs = [('https://www.antutu.com/en/ranking/rank1.htm', 'Android'),
         ('https://www.antutu.com/en/ranking/ios1.htm', 'IOS'),
         ('https://www.antutu.com/en/ranking/ai1.htm', 'AI')]


@pytest.fixture(name='work_book')
def create_book():
    return openpyxl.Workbook()


def test_create_book():
    bk = book.create_book()
    assert type(bk) == openpyxl.workbook.workbook.Workbook


def test_save_book(work_book, tmp_path):
    path = pathlib.Path(tmp_path, 'test.xlsx')
    book.save_book(work_book, path)
    assert path.exists()


@pytest.mark.parametrize('url, platform', pairs)
def test_create_sheet(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    book.create_sheet(work_book, data)
    assert work_book.sheetnames[0] == platform


@pytest.mark.parametrize('url, platform', pairs)
def test_main(url, platform, tmp_path):
    path = pathlib.Path(tmp_path, 'test_main.xlsx')
    data = parser.get_data(parser.parse_text(url), platform)
    book.main(data, path)
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames[0] == platform
    sheet = wb.active
    assert sheet['A1'].value == platform