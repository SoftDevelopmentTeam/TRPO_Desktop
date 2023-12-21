import pytest
import openpyxl
from app.Parser import parser
from app.Excel.Sheets.Styles import styles
from app.Excel.Sheets import sheet


pairs = [('https://www.antutu.com/en/ranking/rank1.htm', 'Android'),
         ('https://www.antutu.com/en/ranking/ios1.htm', 'IOS'),
         ('https://www.antutu.com/en/ranking/ai1.htm', 'AI')]


@pytest.fixture(name='work_book')
def create_book():
    return openpyxl.Workbook()


@pytest.mark.parametrize('url, platform', pairs)
def test_create_sheet(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    work_sheet = sheet.create_sheet(work_book, data)
    assert work_sheet.title == platform


@pytest.mark.parametrize('url, platform', pairs)
def test_set_size(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    work_sheet = sheet.create_sheet(work_book, data)
    sheet.set_size(work_sheet, data)
    assert work_sheet.row_dimensions[1].height == 30


@pytest.mark.parametrize('url, platform', pairs)
def test_set_titles(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    work_sheet = sheet.create_sheet(work_book, data)
    sheet.set_titles(work_sheet, data)
    assert work_sheet['A1'].value == platform


@pytest.mark.parametrize('url, platform', pairs)
def test_set_data(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    work_sheet = sheet.create_sheet(work_book, data)
    sheet.set_titles(work_sheet, data)
    sheet.set_data(work_sheet, data)
    assert work_sheet['A3'].value == data[2][0]


@pytest.mark.parametrize('url, platform', pairs)
def test_set_styles(url, platform, work_book):
    data = parser.get_data(parser.parse_text(url), platform)
    work_sheet = sheet.create_sheet(work_book, data)
    sheet.set_styles(work_sheet, data)
    assert work_sheet['A1'].alignment == styles.alignment()